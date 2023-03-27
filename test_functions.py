from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import ElementClickInterceptedException, NoSuchElementException, InvalidSelectorException, WebDriverException
from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image as PILImage
from selenium.webdriver.common.by import By
from io import BytesIO

import time, openpyxl, os, re, requests

class BasePage(object):
    def __init__(self, driver):
        self.driver = driver

#로그인 페이지 클래스
class LoginPage(BasePage):
    #semdev 로그인 함수
    def dev_login(self, ID, PW):
        self.driver.find_element(By.NAME, "user_id").send_keys(ID)
        self.driver.find_element(By.NAME, "user_pw").send_keys(PW)
        self.driver.find_element(By.CLASS_NAME, "btn").send_keys(Keys.ENTER)
    #p6 로그인 함수
    def p6_login(self, ID, PW):
        self.driver.find_element(By.NAME, "j_username").send_keys(ID)
        self.driver.find_element(By.NAME, "j_password").send_keys(PW)
        self.driver.find_element(By.CLASS_NAME, "coral3-Button").send_keys(Keys.ENTER)

#메인 function 클래스
class MainFunction(BasePage):
    ###########################
    # 1번 산출물: 카피덱 반영 확인
    def copydeck_applied(self, excelName, sheetName):

        # 엑셀 설정
        self.wb = openpyxl.load_workbook(filename=excelName)
        self.ws = self.wb[sheetName]
        self.ws['E2'] = self.driver.current_url

        # 스크린샷 저장을 위한 폴더 설정
        folder_name = "copdydeckText_screenshot"
        current_directory = os.getcwd()
        new_folder_path = os.path.join(current_directory, folder_name)
        if not os.path.exists(new_folder_path):
            os.makedirs(new_folder_path)
        else:
            pass

        time.sleep(3)

        #모든 Q&A 창 펼치기
        ss = "$('.highlights-faq__question-arrow').click()"
        self.driver.execute_script(ss)

        number=0
        pageWholeText = self.driver.find_element(By.XPATH, "/html/body").text

        for i in range(4, 303):
            excelCopydeckText = self.ws['C'+str(i)].value

            if excelCopydeckText is not None and excelCopydeckText in pageWholeText:
                # 텍스트에 개행 존재 시 첫번째 줄의 텍스트만 사용
                first_line_text = excelCopydeckText.split('\n')[0]

                number = number + 1
                self.ws['D'+str(i)] = "PASS"

                if any(char.isdigit() for char in first_line_text):
                    match = re.search(r'^\D+', first_line_text)
                    if match is not None:
                        first_line_text = match.group()
                    else:
                        pass
                else:
                    pass

                try:
                    element = self.driver.find_element(By.XPATH, "//*[contains(text(), \"{}\")]".format(first_line_text))

                    # calculate the x and y coordinates of the element
                    x = element.location['x'] - self.driver.execute_script("return window.innerWidth") / 2
                    y = element.location['y'] - self.driver.execute_script("return window.innerHeight") / 2
                    # scroll to the element
                    self.driver.execute_script("window.scrollTo({0}, {1});".format(x, y))

                    self.driver.execute_script("arguments[0].style.background = 'red'", element)

                    time.sleep(1)

                    screenshot_path = os.path.join(new_folder_path, str(number)+".png")
                    self.driver.save_screenshot(screenshot_path)

                    textImage = ExcelImage(screenshot_path)
                    textImage.height = 500
                    textImage.width = 900
                    self.ws.add_image(textImage, 'E'+str(i))

                    self.driver.execute_script("arguments[0].style.background = ''", element)
                    time.sleep(0.5)
                except InvalidSelectorException as e:
                    print("InvalidSelector Exception occurred: {}".format(str(e)))
                except NoSuchElementException as e:
                    print("NoSuchElement Exception occurred: {}".format(str(e)))
            elif excelCopydeckText is not None and excelCopydeckText not in pageWholeText:
                self.ws['D'+str(i)] = "FAIL"
            elif excelCopydeckText is None:
                self.ws['D'+str(i)] = "None"
            else:
                pass
        
        # 파일저장
        self.wb.save(filename=excelName)

    ###########################
    # 2번 산출물: 하이라이트 페이지 리포트
    def page_report(self, excelName, sheetName, waitTime):
        
        # 엑셀 설정
        self.wb = openpyxl.load_workbook(filename=excelName)
        self.ws = self.wb[sheetName]

        # 스크린샷 저장을 위한 폴더 설정
        folder_name = "elements_screenshot"
        current_directory = os.getcwd()
        new_folder_path = os.path.join(current_directory, folder_name)
        if not os.path.exists(new_folder_path):
            os.makedirs(new_folder_path)
        else:
            pass

        self.ws['C3'] = self.driver.current_url

        # ID 속성이 비어있지 않은 모든 element를 찾고, 특정 ID들과 ID값들을 제외
        elements_with_id = self.driver.find_elements(By.XPATH,"//*[@id and not(@id='wrap') and not(@id='accessibility-navigation') and not(@id='header') and not(@id='gnb') and not(@id='g-products') and not(@id='g-campaigns') and not(@id='g-event') and not(@id='g-apps') and not(@id='cseSearchForm') and not(@id='cseOpenButton') and not(@id='addsearch-body') and not(@id='contents') and not(@id='subnav') and not(@id='faq_item1') and not(@id='faq_item2') and not(@id='faq_item3') and not(@id='faq_item4') and not(@id='faq_item5') and not(@id='faq_item6') and not(@id='faq_item7') and not(@id='faq_item8') and not(@id='desc-section') and not(@id='footer') and not(@id='accessibility-contrast') and not(@id='bandwidth-control') and not(@id='terms-and-conditions') and not(@id='footer-sitemap') and not(@id='modal-layer-popup') and not(@id='color-lavender') and not(@id='color-cream') and not(@id='color-phantom-black') and not(@id='color-green') and not(@id='')]")

        #번호 카운트를 위한 변수
        numberForCount = 0

        # ID를 가지고 있는 모든 element에 접근
        for element in elements_with_id:
            numberForCount = numberForCount + 1
            elementIdValue = element.get_attribute('id')
            # element ID값 엑셀에 쓰기
            self.ws["B"+str(numberForCount+4)] = elementIdValue

            # 해당 element로 이동
            self.driver.execute_script("arguments[0].scrollIntoView();", element)
            time.sleep(waitTime)
            # 100픽셀 만큼 스크롤업
            self.driver.execute_script("window.scrollBy(0,-100);")
            time.sleep(waitTime)
            # 스크린 샷 경로 및 이미지명 지정
            screenshot_path = os.path.join(new_folder_path, str(elementIdValue)+".png")
            # 스크린 샷
            try:
                element.screenshot(screenshot_path)
                elementImage = ExcelImage(screenshot_path)
                elementImage.height = 475
                elementImage.width = 600
                self.ws.add_image(elementImage, 'G'+str(numberForCount+4))
            except WebDriverException:
                self.ws["G"+str(numberForCount+4)] = "Can't take screenshot"
            
            # 클래스 이름이 blind인 element들을 각 element에서 찾기
            blind_elems = element.find_elements(By.CLASS_NAME, "blind")
            # 각 element text 변수
            element_text = element.text

            # 각 element에 이미지가 있는지 확인
            img_elems = element.find_elements(By.TAG_NAME, "img")
            if img_elems:
                all_values_str = ""
                all_alt_str = ""

                for i, img_elem in enumerate(img_elems):
                    img_src_pc = img_elem.get_attribute("data-src-pc")
                    img_src_tablet = img_elem.get_attribute("data-src-tablet")
                    img_src_mobile = img_elem.get_attribute("data-src-mobile")
                    img_alt = img_elem.get_attribute("alt")

                    values_list = []
                    alt_list = []

                    if img_src_pc:
                        values_list.append("Image src-pc: " + str(img_src_pc))
                    if img_src_tablet:
                        values_list.append("Image src-tablet: " + str(img_src_tablet))
                    if img_src_mobile:
                        values_list.append("Image src-mobile: " + str(img_src_mobile))
                    if img_alt:
                        alt_list.append("Image alt: " + str(img_alt))
                    
                    values_str = "\n".join(values_list)
                    all_values_str += values_str + "\n"
                    alt_str = "\n".join(alt_list)
                    all_alt_str += alt_str + "\n"
                
                if all_values_str == " ":
                    self.ws['C'+str(numberForCount+4)] = "None"
                else:
                    self.ws['C'+str(numberForCount+4)] = all_values_str.strip()
                if all_alt_str == " ":
                    self.ws['D'+str(numberForCount+4)] = "None"
                else:
                    self.ws['D'+str(numberForCount+4)] = all_alt_str.strip()
            
            if blind_elems:
                all_blind_str = ""

                for blind_elem in blind_elems:
                    blind_text = blind_elem.get_attribute("innerHTML")
                    blind_list = []
                    if blind_text:
                        blind_list.append("Blind text: " + str(blind_text))
                    
                    blinds_str = "\n".join(blind_list)
                    all_blind_str += blinds_str + "\n"

                    element_text = element_text.replace(blind_elem.text, "")

                if all_blind_str == " ":
                    self.ws['E'+str(numberForCount+4)] = "None"
                else:
                    self.ws['E'+str(numberForCount+4)] = all_blind_str.strip()
            
            if element_text == "":
                self.ws['F'+str(numberForCount+4)] = ""
            else:
                self.ws['F'+str(numberForCount+4)] = element_text
        # 파일저장
        self.wb.save(filename=excelName)

    ###########################
    # 3번 산출물: 악세사리 페이지 리포트
    # kv_visual 
    def kv_visual_src_img(self, idx, acc, list_target, list_route, all_counts, driver, all_images):
        if acc.find_elements(By.TAG_NAME,"source"):
            source_image = acc.find_elements(By.TAG_NAME,"source")
            for source in source_image:
                list_target.append(source.tag_name) # source 태그 이름
                list_route.append(source.get_attribute("srcset")) # source 값
            all_counts.append(len(list_target))

        data_image = acc.find_element(By.TAG_NAME,"img")

        list_target.append("data-src-pc") # 이미지 태그 이름
        list_route.append(data_image.get_attribute("data-src-pc")) # 값
        list_target.append("data-src-tablet") # 이미지 태그 이름
        list_route.append(data_image.get_attribute("data-src-tablet"))# 값
        list_target.append("data-src-mobile") # 이미지 태그 이름
        list_route.append(data_image.get_attribute("data-src-mobile"))# 값
        list_target.append("alt") # 이미지 태그 이름
        list_route.append(data_image.get_attribute("alt"))# 값

        if idx == 0 :
            time.sleep(0.3)
            size = data_image.size
            location = data_image.location

            left = location['x']
            top = location['y']
            right = left + size['width']
            bottom = top + size['height']
            time.sleep(0.3)
            screenshot = driver.get_screenshot_as_png()
            
            image = PILImage.open(BytesIO(screenshot))
            image = image.crop((left, top, right, bottom))


            if not os.path.exists("acc_screenshot"):
                os.makedirs("acc_screenshot")

            folder_name = "acc_screenshot"
            current_directory = os.getcwd()
            screenshot_path = os.path.join(current_directory,folder_name)
            image.save(os.path.join(screenshot_path,"kv.png"))
            all_images.append(os.path.join(screenshot_path,"kv.png"))
            print("save_kv_scr_alt_img")

        elif idx > 0 :
            element = driver.find_elements(By.CSS_SELECTOR,"figure.accessories__visual-image")
            time.sleep(0.3)
            size = element[idx-1].size
            location = element[idx-1].location

            left = location['x']
            top = location['y'] - 96
            right = left + size['width']
            bottom = size['height'] 
                
            driver.execute_script(f"window.scrollTo({left},{top})")
                
            time.sleep(0.5)

            screenshot = driver.get_screenshot_as_png()
            image = PILImage.open(BytesIO(screenshot))
            image = image.crop((left, 96, right, bottom + 96)) 

            folder_name = "acc_screenshot"
            current_directory = os.getcwd()
            screenshot_path = os.path.join(current_directory,folder_name)
            image.save(os.path.join(screenshot_path, "visual" + str(idx) + ".png"))
            all_images.append(os.path.join(screenshot_path ,"visual" + str(idx) + ".png"))
            print("save_visual" + str(idx) + "_scr_alt_img")
        all_counts.append(len(list_target))
    # 제품
    def product_acc_item_scr_img_alt(self,product_counts,idx,acc_product_count,product_option,product_color,counts,click_count,color,list_target,list_route,all_counts,driver,img,all_images):
        idx = idx - 1
        product = product_counts[idx]
        product_option_counts = acc_product_count[idx].find_elements(By.CSS_SELECTOR,"ul.accessories__product-option-list") # 제품 옵션 버튼 

        for i in range(0,product): # 제품 수 만큼 반복  
            product_item = acc_product_count[idx].find_elements(By.CSS_SELECTOR,"li.accessories__product-item")
            product_item_name = product_item[i].find_element(By.CSS_SELECTOR,"div.accessories__product-copy > h3").text # 제품 이름 추출

            if product_option[idx] > 0:
                    product_option_count = product_option_counts[i].find_elements(By.CSS_SELECTOR,"li.accessories__product-option-item")
                                
            if click_count == len(product_color):
                break
                                
            colorchip_click = product_color[click_count].find_elements(By.CSS_SELECTOR,"li.accessories__colorchip-item")
                # 컬러칩 클릭할 요소 추출
            count = len(colorchip_click)
            for j in range(0,count):
                                    
                if product_option[idx] > 0 and j >= (int(count/2)):
                    num = len(product_option_count) - 1
                    product_option_count[num].click()
                    product_name = product_option_count[num].find_element(By.CSS_SELECTOR,"button.accessories__product-option-btn").text # 제품 기종 추출

                elif product_option[idx] == 0 :
                    product_name = ""
                else : 
                    product_name = product_option_count[0].find_element(By.CSS_SELECTOR,"button.accessories__product-option-btn").text #제품 기종 추출

                colorchip_click[j].click()
                product_item_color = product_item[i].find_element(By.CSS_SELECTOR,"span.accessories__product-current").text # 제품 색 추출

                a = product_color[color].find_element(By.TAG_NAME,"img")
                list_target.append("data-src-pc") # 이미지 태그 이름
                list_route.append(a.get_attribute("data-src-pc")) # 값
                list_target.append("data-src-tablet") # 이미지 태그 이름
                list_route.append(a.get_attribute("data-src-tablet"))# 값
                list_target.append("data-src-mobile") # 이미지 태그 이름
                list_route.append(a.get_attribute("data-src-mobile"))# 값
                list_target.append("alt") # 이미지 태그 이름
                list_route.append(a.get_attribute("alt"))# 값    
                time.sleep(0.3)
                                    
                all_counts.append(len(list_target))

                time.sleep(0.3)
                size = product_item[i].size
                location = product_item[i].location

                left = location['x']
                top = location['y'] - 96
                right = left + size['width']
                bottom = size['height'] 
                            
                driver.execute_script(f"window.scrollTo({left},{top})")
                            
                time.sleep(0.5)

                screenshot = driver.get_screenshot_as_png()
                image = PILImage.open(BytesIO(screenshot))
                image = image.crop((left, 96, right, bottom + 96)) 

                folder_name = "acc_screenshot"
                current_directory = os.getcwd()
                screenshot_path = os.path.join(current_directory,folder_name)
                image.save(os.path.join(screenshot_path, product_item_name +" "+ product_name +" "+ product_item_color + ".png"))
                all_images.append(os.path.join(screenshot_path, product_item_name +" "+ product_name +" "+ product_item_color + ".png"))
                img = img + 1

            click_count = click_count + 1
            color = color + 1

        time.sleep(0.3)
        counts = counts + 1
        if idx == 0:
            print("save_cases_src_alt_img")
        elif idx == 1:
            print("save_chargers_src_alt_img")
        elif idx == 2:
            print("save_buds_src_alt_img")
        elif idx == 3:
            print("save_watch_src_alt_img")

        return color, counts, click_count
    # 하단 배너
    def bottom_banner(self,driver, list_target,list_route,all_counts,bann,all_images):
        bann = 0
        acc_banner = driver.find_elements(By.CSS_SELECTOR,"div.common-banner__container")

        for banner in acc_banner:
            banner_img = banner.find_element(By.TAG_NAME,"img")
            list_target.append("data-src-pc") # 이미지 태그 이름
            list_route.append(banner_img.get_attribute("data-src-pc")) # 값
            list_target.append("data-src-tablet") # 이미지 태그 이름
            list_route.append(banner_img.get_attribute("data-src-tablet"))# 값
            list_target.append("data-src-mobile") # 이미지 태그 이름
            list_route.append(banner_img.get_attribute("data-src-mobile"))# 값
            list_target.append("alt") # 이미지 태그 이름
            list_route.append(banner_img.get_attribute("alt"))# 값

            all_counts.append(len(list_target))
            
            banner_inner_img = driver.find_elements(By.CSS_SELECTOR,"div.common-banner__item-inner")  
            time.sleep(0.3)
            size = banner_inner_img[bann].size
            location = banner_inner_img[bann].location

            left = location['x']
            top = location['y'] - 96
            right = left + size['width']
            bottom = size['height'] 
                
            driver.execute_script(f"window.scrollTo({left},{top})")
                
            time.sleep(0.5)

            screenshot = driver.get_screenshot_as_png()
            image = PILImage.open(BytesIO(screenshot))
            image = image.crop((left, 96, right, bottom + 96)) 

            folder_name = "acc_screenshot"
            current_directory = os.getcwd()
            screenshot_path = os.path.join(current_directory,folder_name)
            image.save(os.path.join(screenshot_path,"Bottom_banner"+str(bann)+".png"))
            all_images.append(os.path.join(screenshot_path,"Bottom_banner"+str(bann)+".png"))

            bann = bann + 1
            print("save_bottom_banner"+str(bann)+"_src_alt_img")
            # inners = driver.find_elements(By.CSS_SELECTOR,"div.common-banner__text-inner")
            if acc_banner[1] == banner:
                icon = driver.find_element(By.CSS_SELECTOR,"div.common-banner__ar-icon")
                icon = icon.find_element(By.TAG_NAME,"img")
                list_target.append("data-src-pc") # 이미지 태그 이름
                list_route.append(icon.get_attribute("data-src-pc")) # 값
                list_target.append("alt") # 이미지 태그 이름
                list_route.append(icon.get_attribute("alt"))# 값

                qr = driver.find_element(By.CSS_SELECTOR,"div.common-banner__ar-qr")
                qr = qr.find_element(By.TAG_NAME,"img")
                list_target.append("data-src-pc") # 이미지 태그 이름
                list_route.append(qr.get_attribute("data-src-pc")) # 값
                list_target.append("alt") # 이미지 태그 이름
                list_route.append(qr.get_attribute("alt"))# 값
        return bann
    # 파일 생성
    def make_file(self,list_number,list_target,list_route,all_images, all_counts):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["number","대상", "이미지 경로","응답코드","이미지"])

        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 8
        ws.column_dimensions['C'].width = 80
        ws.column_dimensions['D'].width = 8
        ws.column_dimensions['E'].width = 80

        i = 0
        c = 1
        # a = 1

        for i in range(0,len(list_target)):
            list_number.append(i+1)
            ws.append([list_number[i],list_target[i],list_route[i]])
            
            # 이미지 로드 requset 응답 추가
            if list_route[i] != "" and (list_route[i] is not None) :
                a = i + 1
                res = requests.get("https://www.samsung.com/" + list_route[i]) 
                time.sleep(0.5)
                if res.status_code == requests.codes.ok:
                    text = "Pass"
                    # 원하는 셀 선택
                    cell = ws.cell(row=a+1, column=4)
                    # 셀에 텍스트 추가
                    cell.value = text

            if c <= len(all_images) and i == (all_counts[c])-3:

                img = ExcelImage(all_images[c-1])
                time.sleep(0.3)
                ws.row_dimensions[i+1].height = 300

                if all_counts[c] == 7 or all_counts[c] == 11 or all_counts[c] == 175 or all_counts[c] == 199 or all_counts[c] == 235 or all_counts[c] == 279 or all_counts[c] == 283 or all_counts[c] == 291:
                    ws.row_dimensions[i+1].height = 100
                    img.width = 300
                    img.height = 100
                else : 
                    ws.row_dimensions[i+1].height = 300
                    img.width = 100
                    img.height = 300
                # 이미지를 첨부할 셀 지정
                cell = ws.cell(row=i+1, column=5)
                # 이미지를 셀에 첨부
                ws.add_image(img, cell.coordinate)
                c = c+1
        wb.save('accessories.xlsx')

        if len(list_number) == len(list_target):
            print("file create complete")

    ###########################
    # 4번 산출물: 각주 자동화 리포트
    # 4-1번: 최하단 디스클레이머 카피덱 반영 확인
    def copydeck_disclaimerText_applied(self, excelName, sheetName):
        # 엑셀 설정
        self.wb = openpyxl.load_workbook(filename=excelName)
        self.ws = self.wb[sheetName]
        #최하단 디스클레이머 영역 번호가 몇번까지 있는지 확인
        diclaimernumber = self.driver.find_elements(By. CLASS_NAME, "click_sup")
        maxDisclaimerValue = 0

        for x in diclaimernumber:
            numberContext = x.get_attribute("textContent")
            #가장 큰 각주 번호 찾기
            if int(numberContext) > int(maxDisclaimerValue):
                maxDisclaimerValue = numberContext
            else:
                maxDisclaimerValue = maxDisclaimerValue
        
        for i in range(1, int(maxDisclaimerValue)+1):
            #data-sup 활용하여 최하단 disclaimer element 찾기
            a = self.driver.find_element(By.XPATH, "//*[@data-sup='"+str(i)+"']")
            #최하단 disclaimer element text 가져오기
            disclaimercontext = a.get_attribute("textContent")
            excelcontext = self.ws['C'+str(i+3)].value

            #엑셀파일에 최하단 disclaimer element text 값 쓰기
            self.ws['D'+str(i+3)] = disclaimercontext

        # 파일저장
        self.wb.save(filename=excelName)
    
    # 4-2번: 본문 각주 번호 클릭하여 최하단 디스클레이머 영역으로 이동 확인
    def body_disclaimerNumber_click(self, excelName, sheetName, clickOption):

        # 스크린샷 저장을 위한 폴더 설정
        folder_name = "afterClick_screenshot"
        # get the current working directory
        current_directory = os.getcwd()
        # combine the current directory and folder name to create the path
        new_folder_path = os.path.join(current_directory, folder_name)
        # check if the folder already exists
        if not os.path.exists(new_folder_path):
            # if it doesn't exist, create the folder
            os.makedirs(new_folder_path)
        else:
            pass

        # 엑셀 설정
        self.wb = openpyxl.load_workbook(filename=excelName)
        self.ws = self.wb[sheetName]

        # 본문 각주번호 모두 찾기
        bodyText_disclaimerNumber = self.driver.find_elements(By. CLASS_NAME, "click_sup")
        #각주 합계를 구하기 위한 변수
        totalNumDisclaimer = 0
        #모든 Q&A 창 펼치기
        ss = "$('.highlights-faq__question-arrow').click()"
        self.driver.execute_script(ss)

        #본문에 모든 각주 클릭하여 이동된 화면 스크린샷 저장
        for x in bodyText_disclaimerNumber:
            #1번 입력 시
            if clickOption == "1":
                #해당 element로 스크롤
                self.driver.execute_script("arguments[0].scrollIntoView();", x)
                time.sleep(2)
                #클릭
                self.driver.execute_script("arguments[0].click();", x)
            #2번 입력 시
            elif clickOption == "2":
                #ENTER키를 사용하여 element 클릭
                x.send_keys(Keys.ENTER)
                time.sleep(2)
            
            time.sleep(1)

            #해당 element text 가져오기
            text = x.get_attribute("textContent")
            #현재까지 클릭한 각주 갯수 계산
            totalNumDisclaimer = totalNumDisclaimer + 1
            # 총 각주 갯수로 스크린샷을 특정 폴더에 저장
            screenshot_path = os.path.join(new_folder_path, "after_" + str(totalNumDisclaimer)+".png")
            self.driver.save_screenshot(screenshot_path)

            #엑셀 파일에 입력
            #Number
            self.ws['G'+str(totalNumDisclaimer+3)] = totalNumDisclaimer
            #각주번호
            self.ws['H'+str(totalNumDisclaimer+3)] = text

            afterImg = ExcelImage(screenshot_path)
            afterImg.height = 500
            afterImg.width = 900
            self.ws.add_image(afterImg, 'I'+str(totalNumDisclaimer+3))

            time.sleep(2)
        
        # 파일저장
        self.wb.save(filename=excelName)

    # 4-3번: 55개국 최하단 각주번호 순서 확인
    def bottomDisclaimerNumber_order_check(self, fontUrl, backUrl, excelName, sheetName):
        # 엑셀 설정
        self.wb = openpyxl.load_workbook(filename=excelName)
        self.ws = self.wb[sheetName]

        options = webdriver.ChromeOptions()
        options.add_experimental_option("excludeSwitches", ["enable-logging"])
        options.add_argument('headless')
        options.add_argument("user-agent=Mozilla/5.0 (Macintosh; Intel Mac OS X 10_12_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/61.0.3163.100 Safari/537.36")

        # C3:BE3 범위에 셀에 접근
        for column in self.ws.iter_cols(min_row=3, max_row=3, min_col=3, max_col=57):
            # 각 column에 셀에 접근
            for cell in column:
                # 셀 값 가져오기
                cell_value = cell.value
                url = fontUrl + cell_value + backUrl
                #해당 url로 이동
                driver = webdriver.Chrome(options=options)
                driver.get(url)

                normalNumber = 0

                try:
                    # 최하단 각주 영역 가져오기
                    bottomDisclaimerSection = driver.find_element(By.TAG_NAME, "ol")
                    # 최하단 각주 번호 가져오기
                    eachDisclaimerNumber = bottomDisclaimerSection.find_elements(By.CLASS_NAME, "common-bottom-disclaimer__list-item")
                    # 각주 번호가 1번부터 오름차순으로 되어있는지 확인
                    for n in eachDisclaimerNumber:
                        normalNumber = normalNumber + 1
                        if n.is_displayed:
                            # 각 각주 번호의 "data-sup" 값 가져오기
                            bottomDisclaimerNumber = n.get_attribute("data-sup")
                            # "data-sup" 값 엑셀에 입력
                            self.ws.cell(row= cell.row + normalNumber, column=cell.column, value=int(bottomDisclaimerNumber))
                        else:
                            pass
                    driver.close()
                
                except NoSuchElementException:
                    pass

        # 파일저장
        self.wb.save(filename=excelName)

    ###########################
    # 5번 산출물: 태깅 자동화 리포트
    def tagging_automation_report(self, excelName, sheetName, waitTime):
        
        # 윈도우 사이즈 500x900으로 변경
        self.driver.set_window_size(500, 900)

        # 엑셀 설정
        self.wb = openpyxl.load_workbook(filename=excelName)
        self.ws = self.wb[sheetName]
        self.ws['E2'] = self.driver.current_url

        # 스크린샷 저장을 위한 폴더 설정
        folder_name = "tagging_screenshot"
        current_directory = os.getcwd()
        new_folder_path = os.path.join(current_directory, folder_name)
        if not os.path.exists(new_folder_path):
            os.makedirs(new_folder_path)
        else:
            pass

        time.sleep(waitTime)

        number = 0
        #페이지 전체 source 코드 가져오기
        html_source = self.driver.page_source

        for i in range(4, 303):
            # 엑셀에서 태깅 코드 가져오기
            excelTaggingCode = self.ws['C'+str(i)].value
            if excelTaggingCode is not None and excelTaggingCode in html_source:
                number = number + 1
                self.ws['G'+str(i)] = "PASS"

                # 태깅 형식 변경
                attribute_regex = r'(\S+)="([^"]*)"'
                matches = re.findall(attribute_regex, excelTaggingCode)
                desired_text = '[' + ']['.join([f'{attr}="{value}"' for attr, value in matches]) + ']'

                try:
                    # 해당 태깅의 element 찾기
                    element = self.driver.find_element(By.CSS_SELECTOR, f'{str(desired_text)}')

                    # 태깅 정보 엑셀에 쓰기
                    self.ws['D'+str(i)] = element.tag_name
                    self.ws['E'+str(i)] = str(element.is_displayed())
                    self.ws['F'+str(i)] = element.text

                    # element의 x 와 y 좌표 구하기
                    x = element.location['x'] - self.driver.execute_script("return window.innerWidth") / 2
                    y = element.location['y'] - self.driver.execute_script("return window.innerHeight") / 2
                    # element로 스크롤
                    self.driver.execute_script("window.scrollTo({0}, {1});".format(x, y))

                    # 해당 태깅 버튼에 빨간색 강조 처리
                    self.driver.execute_script("arguments[0].style.background = 'red'", element)

                    time.sleep(waitTime)

                    screenshot_path = os.path.join(new_folder_path, str(number)+".png")

                    if str(element.is_displayed()) == "True":
                        self.driver.save_screenshot(screenshot_path)
                        tagImage = ExcelImage(screenshot_path)
                        tagImage.height = 500
                        tagImage.width = 900
                        self.ws.add_image(tagImage, 'H' +str(i))
                    else:
                        pass

                    #강조처리 했던것 초기화
                    self.driver.execute_script("arguments[0].style.background = ''", element)
                    time.sleep(waitTime)
                except InvalidSelectorException as e:
                    print("InvalidSelector Exception occurred: {}".format(str(e)))
                except NoSuchElementException as e:
                    print("NoSuchElement Exception occurred: {}".format(str(e)))
            elif excelTaggingCode is not None and excelTaggingCode not in html_source:
                self.ws['G'+str(i)] = "FAIL"
            elif excelTaggingCode is None:
                self.ws['G'+str(i)] = "None"
            else:
                pass
        # 파일저장
        self.wb.save(filename=excelName)
