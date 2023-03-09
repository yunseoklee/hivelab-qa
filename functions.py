from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from locators import *
from selenium.common.exceptions import ElementClickInterceptedException
from selenium.common.exceptions import NoSuchElementException
from openpyxl.drawing.image import Image

import time, openpyxl, os


class BasePage(object):
    def __init__(self, driver):
        self.driver = driver
        

#로그인 페이지 클래스
class LoginPage(BasePage):
    #semdev 로그인 함수
    def dev_login(self, ID, PW):
        self.driver.find_element(*DevLoginPageLocators.ID_INPUT).send_keys(ID)
        self.driver.find_element(*DevLoginPageLocators.PW_INPUT).send_keys(PW)
        self.driver.find_element(*DevLoginPageLocators.LOGIN_BTN).send_keys(Keys.ENTER)
    #p6 로그인 함수
    def p6_login(self, ID, PW):
        self.driver.find_element(*P6LoginPageLocators.ID_INPUT).send_keys(ID)
        self.driver.find_element(*P6LoginPageLocators.PW_INPUT).send_keys(PW)
        self.driver.find_element(*P6LoginPageLocators.LOGIN_BTN).send_keys(Keys.ENTER)


#Highlights 페이지 클래스
class HighlightsPage(BasePage):
    #컬러칩 클릭
    def colorchip_click_screenshot(self, waitTime):
        # 컬러칩 element 찾기
        colorchips = self.driver.find_elements(*HighlightsPageLocators.COLORCHIP)
        # 현재 작업중인 폴더찾기
        cwd = os.getcwd()
        # 스크린샷 저장할 폴더 경로
        screenshot_dir = os.path.join(cwd, "highlight_colorchip_screenshot")
        # 스크린샷 저장할 폴더 미존재 시 만들기
        if not os.path.exists(screenshot_dir):
            os.makedirs(screenshot_dir)
        # 모든 컬러칩에 순차적으로 접근
        for x in colorchips:
            #컬러칩 클릭
            x.click()
            # 스크린샷 경로 설정
            screenshot_path = os.path.join(screenshot_dir, str(x.text) + ".png")
            time.sleep(waitTime)
            # 스크린샷 할 이미지 element
            screenshot_element = self.driver.find_element(*HighlightsPageLocators.COLORS_SECTION)
            # 스크린샷 하여 스크린샷 경로에 저장
            screenshot_element.screenshot(screenshot_path)
            time.sleep(waitTime)

    # copydeck 텍스트 반영 확인
    def copydeck_text_applied(self, sectionName, cell_range, excelName):
        # 엑셀 설정
        self.wb = openpyxl.load_workbook(filename=excelName)
        self.ws = self.wb.active
        # 섹션 이름을 locator에서 받아와 text 가져오기
        SectionText = self.driver.find_element(*sectionName).text
        # print(SectionText)
        # 셀 범위에서 각 셀 별 값 읽기
        for cell in self.ws[cell_range]:
            # 각 셀의 값
            column_value = cell[0].value
            # 결과 값을 현재셀 우측에 입력
            result_cell = self.ws.cell(row=cell[0].row, column=cell[0].column+1)
            # 셀 값 미존재 시 
            if str(column_value) == "None":
                # 결과셀에 'None' 입력
                result_cell.value = 'None'
            # 셀 값 존재 시
            else:
                # 셀 값이 해당 섹션에 존재 시
                if str(column_value) in SectionText:
                    # 결과셀에 'Pass' 입력
                    result_cell.value = 'Pass'
                # 셀 값이 해당 섹션에 미 존재 시
                else:
                    # 결과셀에 'Fail' 입력
                    result_cell.value = 'Fail'
        # 파일저장
        self.wb.save(filename=excelName)


#Compare 페이지 클래스
class ComparePage(BasePage):
    #컬러칩 클릭
    def colorchip_click_screenshot(self, waitTime):
        # 모든 칼럼 찾기
        columns = self.driver.find_elements(*ComparePageLocators.COLUMN)
        # 모든 칼럼에 순차적으로 접근
        for i, column in enumerate(columns):
            # 디바이스 모델 element
            devices = column.find_elements(*ComparePageLocators.DEVICE)
            # 디바이스 리스트 BTN element
            deviceListBtn = column.find_element(*ComparePageLocators.SELECT_DEVICE_BTN)
            # 현재 작업중인 폴더찾기
            cwd = os.getcwd()
            # 스크린샷 저장할 폴더 경로
            screenshot_dir = os.path.join(cwd, "compare_colorchip_screenshot")
            # 스크린샷 저장할 폴더 미존재 시 만들기
            if not os.path.exists(screenshot_dir):
                os.makedirs(screenshot_dir)
            # 모든 디바이스에 순차적으로 접근
            for d in devices:
                # 디바이스 목록 확장 버튼 클릭
                deviceListBtn.click()
                time.sleep(waitTime)
                # 디바이스 클릭
                d.click()
                # 컬러칩 element
                colorchips = column.find_elements(*ComparePageLocators.COLORCHIP)
                #모든 컬러칩 클릭 후 이미지 저장
                for x in colorchips:
                    try:
                        #컬러칩이 페이지에 보여질 시에만
                        if x.is_displayed():
                            #컬러칩 클릭
                            x.click()
                            #파일명을 위한 변수
                            colorchipLabel = x.find_element(*ComparePageLocators.COLORCHIP_LABEL)
                            fileName = colorchipLabel.get_attribute("for")

                            # set the screenshot path
                            screenshot_path = os.path.join(screenshot_dir, str(fileName) + ".png")
                            time.sleep(waitTime)
                            screenshot_element = column.find_element(*ComparePageLocators.COLORCHIP_IMAGE)
                            screenshot_element.screenshot(screenshot_path)
                            time.sleep(waitTime)
                    except ElementClickInterceptedException:
                        print("Can't click")


#Common 클래스
class Common(BasePage):
    #디스클레이머 카피덱 반영 확인
    def copydeck_disclaimerText_applied(self, excelName, sheetName):
        # 엑셀 설정
        self.wb = openpyxl.load_workbook(filename=excelName)
        self.ws = self.wb[sheetName]
        #최하단 디스클레이머 영역 번호가 몇번까지 있는지 확인
        diclaimernumber = self.driver.find_elements(*CommonLocators.DISCLAIMER_NUMBERS_IN_MAINTEXT)
        maxDisclaimerValue = 0

        for x in diclaimernumber:
            numberContext = x.get_attribute("textContent")
            #가장 큰 각주 번호 찾기
            if int(numberContext) > int(maxDisclaimerValue):
                maxDisclaimerValue = numberContext
            else:
                maxDisclaimerValue = maxDisclaimerValue
        
        for i in range(1, int(maxDisclaimerValue)+1):
            #data-sup 활용하여 최하단 disclaimer element 찾기
            a = self.driver.find_element(By.XPATH, "//*[@data-sup='"+str(i)+"']")
            #최하단 disclaimer element text 가져오기
            disclaimercontext = a.get_attribute("textContent")
            excelcontext = self.ws['C'+str(i+3)].value

            #엑셀파일에 최하단 disclaimer element text 값 쓰기
            self.ws['D'+str(i+3)] = disclaimercontext

        # 파일저장
        self.wb.save(filename=excelName)
    
    #본문 각주 번호 클릭하려 최하단 디스클레이머 영역으로 이동 확인
    def body_disclaimerNumber_click(self, excelName, sheetName, clickOption):
        
        # 스크린샷 저장을 위한 폴더 설정
        folder_name = "afterClick_screenshot"
        # get the current working directory
        current_directory = os.getcwd()
        # combine the current directory and folder name to create the path
        new_folder_path = os.path.join(current_directory, folder_name)
        # check if the folder already exists
        if not os.path.exists(new_folder_path):
            # if it doesn't exist, create the folder
            os.makedirs(new_folder_path)
        else:
            pass

        # 엑셀 설정
        self.wb = openpyxl.load_workbook(filename=excelName)
        self.ws = self.wb[sheetName]

        # 본문 각주번호 모두 찾기
        bodyText_disclaimerNumber = self.driver.find_elements(*CommonLocators.DISCLAIMER_NUMBERS_IN_MAINTEXT)
        #각주 합계를 구하기 위한 변수
        totalNumDisclaimer = 0
        #모든 Q&A 창 펼치기
        ss = "$('.highlights-faq__question-arrow').click()"
        self.driver.execute_script(ss)

        #본문에 모든 각주 클릭하여 이동된 화면 스크린샷 저장
        for x in bodyText_disclaimerNumber:
            #1번 입력 시
            if clickOption == "1":
                #해당 element로 스크롤
                self.driver.execute_script("arguments[0].scrollIntoView();", x)
                time.sleep(2)
                #클릭
                self.driver.execute_script("arguments[0].click();", x)
            #2번 입력 시
            elif clickOption == "2":
                #ENTER키를 사용하여 element 클릭
                x.send_keys(Keys.ENTER)
                time.sleep(2)
            
            time.sleep(1)

            #해당 element text 가져오기
            text = x.get_attribute("textContent")
            #현재까지 클릭한 각주 갯수 계산
            totalNumDisclaimer = totalNumDisclaimer + 1
            # 총 각주 갯수로 스크린샷을 특정 폴더에 저장
            screenshot_path = os.path.join(new_folder_path, "after_" + str(totalNumDisclaimer)+".png")
            self.driver.save_screenshot(screenshot_path)

            #엑셀 파일에 입력
            #Number
            self.ws['G'+str(totalNumDisclaimer+3)] = totalNumDisclaimer
            #각주번호
            self.ws['H'+str(totalNumDisclaimer+3)] = text

            afterImg = Image(screenshot_path)
            afterImg.height = 500
            afterImg.width = 900
            self.ws.add_image(afterImg, 'I'+str(totalNumDisclaimer+3))

            time.sleep(2)
        
        # 파일저장
        self.wb.save(filename=excelName)
    
    # 55개국 최하단 각주 번호 순서대로 출력확인
    def bottomDisclaimerNumber_order_check(self, excelName, sheetName):
        # 엑셀 설정
        self.wb = openpyxl.load_workbook(filename=excelName)
        self.ws = self.wb[sheetName]

        options = webdriver.ChromeOptions()
        options.add_experimental_option("excludeSwitches", ["enable-logging"])
        options.add_argument('headless')
        options.add_argument("user-agent=Mozilla/5.0 (Macintosh; Intel Mac OS X 10_12_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/61.0.3163.100 Safari/537.36")

        # C3:BE3 범위에 셀에 접근
        for column in self.ws.iter_cols(min_row=3, max_row=3, min_col=3, max_col=57):
            # 각 column에 셀에 접근
            for cell in column:
                # 셀 값 가져오기
                cell_value = cell.value
                url = "https://www.samsung.com/" + str(cell_value) + "/smartphones/galaxy-s23-ultra/"
                #해당 url로 이동
                driver = webdriver.Chrome(options=options)
                driver.get(url)

                normalNumber = 0

                try:
                    # 최하단 각주 영역 가져오기
                    bottomDisclaimerSection = driver.find_element(*CommonLocators.BOTTOM_DISCLAIMER_SECTION)
                    # 최하단 각주 번호 가져오기
                    eachDisclaimerNumber = bottomDisclaimerSection.find_elements(*CommonLocators.EACH_BOTTOM_DISCLAIMER_ELEMENT)
                    # 각주 번호가 1번부터 오름차순으로 되어있는지 확인
                    for n in eachDisclaimerNumber:
                        normalNumber = normalNumber + 1
                        if n.is_displayed:
                            # 각 각주 번호의 "data-sup" 값 가져오기
                            bottomDisclaimerNumber = n.get_attribute("data-sup")
                            # "data-sup" 값 엑셀에 입력
                            self.ws.cell(row= cell.row + normalNumber, column=cell.column, value=int(bottomDisclaimerNumber))
                        else:
                            pass
                    driver.close()
                
                except NoSuchElementException:
                    pass

        # 파일저장
        self.wb.save(filename=excelName)













