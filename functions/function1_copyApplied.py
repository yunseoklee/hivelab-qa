from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl.drawing.image import Image as ExcelImage
from selenium.common.exceptions import InvalidSelectorException, NoSuchElementException
import openpyxl, os, time, re

class BasePage(object):
    def __init__(self, driver):
        self.driver = driver

# 1번 산출물: 카피덱 반영 확인
#메인 function 클래스
class MainFunction(BasePage):
    def copydeck_applied(self, excelName, sheetName):

        # 엑셀 설정
        self.wb = openpyxl.load_workbook(filename=excelName)
        self.ws = self.wb[sheetName]
        self.ws['E2'] = self.driver.current_url

        # 스크린샷 저장을 위한 폴더 설정
        folder_name = "copdydeckText_screenshot"
        current_directory = os.getcwd()
        new_folder_path = os.path.join(current_directory, folder_name)
        if not os.path.exists(new_folder_path):
            os.makedirs(new_folder_path)
        else:
            pass

        time.sleep(3)

        #모든 Q&A 창 펼치기
        ss = "$('.highlights-faq__question-arrow').click()"
        self.driver.execute_script(ss)

        number=0
        pageWholeText = self.driver.find_element(By.XPATH, "/html/body").text

        for i in range(4, 303):
            excelCopydeckText = self.ws['C'+str(i)].value

            if excelCopydeckText is not None and excelCopydeckText in pageWholeText:
                # 텍스트에 개행 존재 시 첫번째 줄의 텍스트만 사용
                first_line_text = excelCopydeckText.split('\n')[0]

                number = number + 1
                self.ws['D'+str(i)] = "PASS"

                if any(char.isdigit() for char in first_line_text):
                    match = re.search(r'^\D+', first_line_text)
                    if match is not None:
                        first_line_text = match.group()
                    else:
                        pass
                else:
                    pass

                try:
                    element = self.driver.find_element(By.XPATH, "//*[contains(text(), \"{}\")]".format(first_line_text))

                    # calculate the x and y coordinates of the element
                    x = element.location['x'] - self.driver.execute_script("return window.innerWidth") / 2
                    y = element.location['y'] - self.driver.execute_script("return window.innerHeight") / 2
                    # scroll to the element
                    self.driver.execute_script("window.scrollTo({0}, {1});".format(x, y))

                    self.driver.execute_script("arguments[0].style.background = 'red'", element)

                    time.sleep(1)

                    screenshot_path = os.path.join(new_folder_path, str(number)+".png")
                    self.driver.save_screenshot(screenshot_path)

                    textImage = ExcelImage(screenshot_path)
                    textImage.height = 500
                    textImage.width = 900
                    self.ws.add_image(textImage, 'E'+str(i))

                    self.driver.execute_script("arguments[0].style.background = ''", element)
                    time.sleep(0.5)
                except InvalidSelectorException as e:
                    print("InvalidSelector Exception occurred: {}".format(str(e)))
                except NoSuchElementException as e:
                    print("NoSuchElement Exception occurred: {}".format(str(e)))
            elif excelCopydeckText is not None and excelCopydeckText not in pageWholeText:
                self.ws['D'+str(i)] = "FAIL"
            elif excelCopydeckText is None:
                self.ws['D'+str(i)] = "None"
            else:
                pass
        
        # 파일저장
        self.wb.save(filename=excelName)
