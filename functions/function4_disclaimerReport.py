from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from openpyxl.drawing.image import Image as ExcelImage
from selenium.common.exceptions import NoSuchElementException
import openpyxl, os, time

class BasePage(object):
    def __init__(self, driver):
        self.driver = driver

#메인 function 클래스
class MainFunction(BasePage):
    # 4-1번: 최하단 디스클레이머 카피덱 반영 확인
    def copydeck_disclaimerText_applied(self, excelName, sheetName, waitTime):
        # 엑셀 설정
        self.wb = openpyxl.load_workbook(filename=excelName)
        self.ws = self.wb[sheetName]

        time.sleep(waitTime)

        #최하단 디스클레이머 영역 번호가 몇번까지 있는지 확인
        diclaimernumber = self.driver.find_elements(By. CLASS_NAME, "click_sup")
        maxDisclaimerValue = 0

        for x in diclaimernumber:
            numberContext = x.get_attribute("textContent")
            #가장 큰 각주 번호 찾기
            if int(numberContext) > int(maxDisclaimerValue):
                maxDisclaimerValue = numberContext
            else:
                maxDisclaimerValue = maxDisclaimerValue
        
        for i in range(1, int(maxDisclaimerValue)+1):
            #data-sup 활용하여 최하단 disclaimer element 찾기
            a = self.driver.find_element(By.XPATH, "//*[@data-sup='"+str(i)+"']")
            #최하단 disclaimer element text 가져오기
            disclaimercontext = a.get_attribute("textContent")
            excelcontext = self.ws['C'+str(i+3)].value

            #엑셀파일에 최하단 disclaimer element text 값 쓰기
            self.ws['D'+str(i+3)] = disclaimercontext

        # 파일저장
        self.wb.save(filename=excelName)

    # 4-2번: 본문 각주 번호 클릭하여 최하단 디스클레이머 영역으로 이동 확인
    def body_disclaimerNumber_click(self, excelName, sheetName, clickOption, waitTime):

        # 스크린샷 저장을 위한 폴더 설정
        folder_name = "afterClick_screenshot"
        # get the current working directory
        current_directory = os.getcwd()
        # combine the current directory and folder name to create the path
        new_folder_path = os.path.join(current_directory, folder_name)
        # check if the folder already exists
        if not os.path.exists(new_folder_path):
            # if it doesn't exist, create the folder
            os.makedirs(new_folder_path)
        else:
            pass

        # 엑셀 설정
        self.wb = openpyxl.load_workbook(filename=excelName)
        self.ws = self.wb[sheetName]

        time.sleep(waitTime)

        # 본문 각주번호 모두 찾기
        bodyText_disclaimerNumber = self.driver.find_elements(By. CLASS_NAME, "click_sup")
        #각주 합계를 구하기 위한 변수
        totalNumDisclaimer = 0
        # 쿠키 팝업 닫기
        cookie = "$('.truste-custom-samsung-link').click()"
        self.driver.execute_script(cookie)
        #모든 Q&A 창 펼치기
        ss = "$('.highlights-faq__question-arrow').click()"
        self.driver.execute_script(ss)

        #본문에 모든 각주 클릭하여 이동된 화면 스크린샷 저장
        for x in bodyText_disclaimerNumber:
            #1번 입력 시
            if clickOption == "1":
                #해당 element로 스크롤
                self.driver.execute_script("arguments[0].scrollIntoView();", x)
                time.sleep(waitTime)
                #클릭
                self.driver.execute_script("arguments[0].click();", x)
            #2번 입력 시
            elif clickOption == "2":
                #ENTER키를 사용하여 element 클릭
                x.send_keys(Keys.ENTER)
                time.sleep(waitTime)
            
            time.sleep(waitTime)

            #해당 element text 가져오기
            text = x.get_attribute("textContent")
            #현재까지 클릭한 각주 갯수 계산
            totalNumDisclaimer = totalNumDisclaimer + 1
            # 총 각주 갯수로 스크린샷을 특정 폴더에 저장
            screenshot_path = os.path.join(new_folder_path, "after_" + str(totalNumDisclaimer)+".png")
            self.driver.save_screenshot(screenshot_path)

            #엑셀 파일에 입력
            #Number
            self.ws['G'+str(totalNumDisclaimer+3)] = totalNumDisclaimer
            #각주번호
            self.ws['H'+str(totalNumDisclaimer+3)] = text

            afterImg = ExcelImage(screenshot_path)
            afterImg.height = 500
            afterImg.width = 900
            self.ws.add_image(afterImg, 'I'+str(totalNumDisclaimer+3))

            time.sleep(waitTime)
        
        # 파일저장
        self.wb.save(filename=excelName)

    # 4-3번: 55개국 최하단 각주번호 순서 확인
    def bottomDisclaimerNumber_order_check(self, fontUrl, backUrl, excelName, sheetName, waitTime):
        # 엑셀 설정
        self.wb = openpyxl.load_workbook(filename=excelName)
        self.ws = self.wb[sheetName]

        # options = webdriver.ChromeOptions()
        # options.add_experimental_option("excludeSwitches", ["enable-logging"])
        # options.add_argument('headless')
        # options.add_argument("user-agent=Mozilla/5.0 (Macintosh; Intel Mac OS X 10_12_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/61.0.3163.100 Safari/537.36")

        # C3:BE3 범위에 셀에 접근
        for column in self.ws.iter_cols(min_row=3, max_row=3, min_col=3, max_col=57):
            # 각 column에 셀에 접근
            for cell in column:
                # 셀 값 가져오기
                cell_value = cell.value
                url = fontUrl + cell_value + backUrl

                time.sleep(waitTime)
                #해당 url로 이동
                self.driver.get(url)

                normalNumber = 0

                try:
                    # 최하단 각주 영역 가져오기
                    bottomDisclaimerSection = self.driver.find_element(By.TAG_NAME, "ol")
                    # 최하단 각주 번호 가져오기
                    eachDisclaimerNumber = bottomDisclaimerSection.find_elements(By.CLASS_NAME, "common-bottom-disclaimer__list-item")
                    # 각주 번호가 1번부터 오름차순으로 되어있는지 확인
                    for n in eachDisclaimerNumber:
                        normalNumber = normalNumber + 1
                        if n.is_displayed:
                            # 각 각주 번호의 "data-sup" 값 가져오기
                            bottomDisclaimerNumber = n.get_attribute("data-sup")
                            # "data-sup" 값 엑셀에 입력
                            self.ws.cell(row= cell.row + normalNumber, column=cell.column, value=int(bottomDisclaimerNumber))
                        else:
                            pass
                
                except NoSuchElementException as e:
                    print("NoSuchElement Exception occurred: {}".format(str(e)))

        # 파일저장
        self.wb.save(filename=excelName)