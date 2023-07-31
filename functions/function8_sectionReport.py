import openpyxl
from bs4 import BeautifulSoup
from selenium.webdriver.common.by import By
from selenium.common.exceptions import WebDriverException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from urllib.error import URLError, HTTPError
import sys
import time 


class BasePage(object):
    def __init__(self, driver):
        self.driver = driver

class MainFunction(BasePage):

    def sectioncheck(self, excelName, sheetName):

        self.wb = openpyxl.load_workbook(filename=excelName)
        self.ws = self.wb[sheetName]

        num = 3 # n번째 열을 의미

        d12_id_lst = ['wearable-tab-banner']
        # 확인해야할 구간의 피쳐 값을 잡음 (주로 아이디 값이나 클래스 값)
        # 아이디 값 양식 : #id 값  / '#KV', '#s-pen','#notepaper-screen','#book-cover'
        # 클래스 값 양식 : .클래스 값 / .wearable-tab-banner__list

        for a in d12_id_lst : 
            for i in range(2, self.ws.max_row+1) :
                country_url = self.ws['B'+str(i)].value 
                print('총 :', self.ws.max_row -1 ) 
                self.driver.get(country_url)

                wait = WebDriverWait(self.driver,100)
                time.sleep(5) 
                if country_url != self.driver.current_url : # country_url이 현재 driver의 URL과 같지 않은지 확인
                    self.driver.find_element(By.ID, "username").send_keys("haem000") # 아이디 
                    self.driver.find_element(By.ID, "password").send_keys("haem000!2023") # 비번
                    self.driver.find_element(By.XPATH, '//*[@id="submit-button"]').click() # 로그인버튼 클릭
                    WebDriverWait(self.driver,20) 
                time.sleep(10)
                html = self.driver.page_source
                soup = BeautifulSoup(html,'html.parser') 
                num = num+1 # 
                print(country_url,' ',i-1,a,'몇번 ',num)

                wait.until(EC.presence_of_element_located((By.TAG_NAME,"body")))

                try : # 예외 처리를 위해 사용
                    if a == '#KV' : #피쳐 o 배너 o CTA o
                        self.ws.cell(row=1,column=3).value = '#KV'
                        KV_div = soup.select_one(a) #id_cta
                        if KV_div == None : # 피쳐 유무
                            print('KV_div가 없음')
                            self.ws.cell(row=i,column=3).value = "KV 없음"
                        else :
                            KV_cta = KV_div.find(attrs={"class":"wearable-tab-kv__cta--order"})
                            print('KV_CTA',KV_cta )
                            if KV_cta == None : # 배너의 버튼 유무
                                self.ws.cell(row=i,column=3).value = "CTA 없음"
                            else :
                                href = KV_cta.find('a').get('href') # 있으면 링크 출력
                                self.ws.cell(row=i,column=3).value = href
                                self.ws.cell(row=i,column=4).value = KV_cta.text # 저장될 링크 column값에 +1

                    if a == '#s-pen' : #피쳐 o 배너 o CTA o
                        self.ws.cell(row=1,column=5).value = '#s-pen'
                        s_pen_div = soup.select_one(a) #id_cta
                        print(s_pen_div)
                        if s_pen_div == None : # 피쳐 유무
                            print('s_pen_div가 없음')
                            self.ws.cell(row=i,column=5).value = "피쳐없음"
                        else : 
                            s_pen_banner = s_pen_div.find(attrs={"class":"wearable-tab-s-pen__banner"})
                            if s_pen_banner == None : # 배너 유무
                                self.ws.cell(row=i,column=5).value = "배너 없음"
                            else :
                                s_pen_cta = s_pen_banner.find(attrs={"class":"wearable-tab-s-pen__banner-cta"})
                                print('smart_cta',s_pen_cta )
                                if s_pen_cta == None : # 배너의 버튼 유무
                                    self.ws.cell(row=i,column=5).value = "CTA 없음"
                                else : 
                                    href = s_pen_cta.find('a').get('href') # 있으면 링크 출력
                                    self.ws.cell(row=i,column=5).value = href
                                    self.ws.cell(row=i,column=6).value = s_pen_cta.text # 저장될 링크 column값에 +1

                    elif a == '#notepaper-screen' : #피쳐 o 배너 o CTA o
                        self.ws.cell(row=1,column=7).value = '#notepaper-screen'
                        notepaper_screen_div = soup.select_one(a) #id_cta
                        if notepaper_screen_div == None : # 피쳐 유무
                            print('notepaper_screen_div가 없음')
                            self.ws.cell(row=i,column=7).value = "피쳐없음"
                        else : 
                            notepaper_screen_banner = notepaper_screen_div.find(attrs={"class":"wearable-tab-paper__banner"})
                            if notepaper_screen_banner == None : # 배너 유무
                                self.ws.cell(row=i,column=7).value = "배너 없음"
                            else :
                                notepaper_screen_cta = notepaper_screen_banner.find(attrs={"class":"wearable-tab-paper__banner-cta"})
                                print('smart_cta',notepaper_screen_cta )
                                if notepaper_screen_cta == None : # 배너의 버튼 유무
                                    self.ws.cell(row=i,column=7).value = "CTA 없음"
                                else : 
                                    href = notepaper_screen_cta.find('a').get('href') # 있으면 링크 출력
                                    self.ws.cell(row=i,column=7).value = href
                                    self.ws.cell(row=i,column=8).value = notepaper_screen_cta.text # 저장될 링크 column값에 +1

                    elif a == '#book-cover' : #피쳐 o 배너 o CTA o
                        self.ws.cell(row=1,column=9).value = '#book-cover'
                        book_cover_div = soup.select_one(a) #id_cta
                        if book_cover_div == None : # 피쳐 유무
                            print('book_cover_div가 없음')
                            self.ws.cell(row=i,column=9).value = "피쳐없음"
                        else : 
                            book_cover_banner = book_cover_div.find(attrs={"class":"wearable-tab-cover__banner"})
                            print('smart_banner',book_cover_banner )
                            if book_cover_banner == None : # 배너 유무
                                self.ws.cell(row=i,column=9).value = "배너 없음"
                            else :
                                book_cover_cta = book_cover_banner.find(attrs={"class":"wearable-tab-cover__banner-cta"})
                                print('smart_cta',book_cover_cta )
                                if book_cover_cta == None : # 배너의 버튼 유무
                                    self.ws.cell(row=i,column=9).value = "CTA 없음"
                                else : 
                                    href = book_cover_cta.find('a').get('href') # 있으면 링크 출력
                                    self.ws.cell(row=i,column=9).value = href 
                                    self.ws.cell(row=i,column=10).value = book_cover_cta.text # 저장될 링크 column값에 +1

                    if a == 'wearable-tab-banner' :
                        self.ws.cell(row=1,column=12).value = 'wearable-tab-banner'
                        tab_banner = soup.select_one('.wearable-tab-banner__list')
                        # print(tab_banner)
                        if tab_banner == None :
                            self.ws.cell(row=i,column=12).value = "피쳐없음"
                            print("비투비 없음")
                        else : 
                            tab_banner_CTA = tab_banner.find_all('a',attrs={"class":"wearable-tab-common-cta"})
                            print(tab_banner_CTA)
                            for b in range(0,len(tab_banner_CTA)):
                                href = tab_banner_CTA[b].get('href')
                                if href == None : 
                                    self.ws.cell(row=i,column=b+12).value = "확인바람"
                                else : 
                                    self.ws.cell(row=i,column=b+12).value = href
                            b = b+1
                    ##d1,2 faq            
                    else : # 위과정이 안되면 로그인 오류 문구 출력
                        print('로긘오류')
                except WebDriverException: # 페이지 오류 처리
                        print('페이지오류')
                except HTTPError as e : # HTTPError 오류 처리
                        print("ERROR"+str(e.code))
                except URLError as e : # URLError 오류처리
                        print("reason"+str(e.reason)) 
        self.driver.close() 
        result=input("저장할 파일명 : ") # 저장할 파일 이름 입력
        self.wb.save(result+'.xlsx') # 입력한 파일로 저장
        self.wb.close() 
        print("완료")
        sys.exit()
