from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image as PILImage
from selenium.webdriver.common.by import By
from io import BytesIO
import time, openpyxl, os, requests

class BasePage(object):
    def __init__(self, driver):
        self.driver = driver

# 3번 산출물: 악세서리 페이지 리포트
#메인 function 클래스
class MainFunction(BasePage):
    def acc_page_report(self):
        self.driver.maximize_window()
        acc_image = self.driver.find_elements(By.CSS_SELECTOR,"picture")
        product_color = self.driver.find_elements(By.CSS_SELECTOR, "div.accessories__product-contents")
        acc_product_count = self.driver.find_elements(By.CSS_SELECTOR,"div.accessories__product")


        product_counts = []# 제품 구간
        for count in  acc_product_count:
            count = count.find_elements(By.CSS_SELECTOR,"li.accessories__product-item")
            product_counts.append(len(count))

        product_option = []# 제품 구간 별 제품 옵션 구간
        for count in  acc_product_count:
            count = count.find_elements(By.CSS_SELECTOR,"ul.accessories__product-option-list")
            product_option.append(len(count))

        all_counts = [] 


        color = 0
        counts = 0
        click_count = 0

        img = 0 #이미지 수 
        bann = 0

        list_number = [] # 숫자
        list_target = [] # 대상
        list_route = [] # 값
        all_images = [] # 이미지 경로 리스트 

        folder_name = "acc_screenshot"

        if not os.path.exists(folder_name):
            os.makedirs(folder_name)

        current_directory = os.getcwd()
        screenshot_path = os.path.join(current_directory,folder_name)
            

        for idx, acc in enumerate(acc_image):
            if acc.find_elements(By.TAG_NAME,"source"):
                source_image = acc.find_elements(By.TAG_NAME,"source")
                for source in source_image:
                    list_target.append(source.tag_name) # source 태그 이름
                    list_route.append(source.get_attribute("srcset")) # source 값
                all_counts.append(len(list_target))

            data_image = acc.find_element(By.TAG_NAME,"img")

            list_target.append("data-src-pc") # 이미지 태그 이름
            list_route.append(data_image.get_attribute("data-src-pc")) # 값
            list_target.append("data-src-tablet") # 이미지 태그 이름
            list_route.append(data_image.get_attribute("data-src-tablet"))# 값
            list_target.append("data-src-mobile") # 이미지 태그 이름
            list_route.append(data_image.get_attribute("data-src-mobile"))# 값
            list_target.append("alt") # 이미지 태그 이름
            list_route.append(data_image.get_attribute("alt"))# 값

            if idx == 0 :
                time.sleep(0.3)
                size = data_image.size
                location = data_image.location

                left = location['x']
                top = location['y']
                right = left + size['width']
                bottom = top + size['height']
                time.sleep(0.3)
                screenshot = self.driver.get_screenshot_as_png()
                    
                image = PILImage.open(BytesIO(screenshot))
                image = image.crop((left, top, right, bottom))

                image.save(os.path.join(screenshot_path,"kv.png"))
                all_images.append(os.path.join(screenshot_path,"kv.png"))
                print("save_kv_scr_alt_img")

            elif idx > 0 :
                element = self.driver.find_elements(By.CSS_SELECTOR,"figure.accessories__visual-image")
                time.sleep(0.3)
                size = element[idx-1].size
                location = element[idx-1].location

                left = location['x']
                top = location['y'] - 96
                right = left + size['width']
                bottom = size['height'] 
                        
                self.driver.execute_script(f"window.scrollTo({left},{top})")
                        
                time.sleep(0.5)

                screenshot = self.driver.get_screenshot_as_png()
                image = PILImage.open(BytesIO(screenshot))
                image = image.crop((left, 96, right, bottom + 96)) 

                image.save(os.path.join(screenshot_path, "visual" + str(idx) + ".png"))
                all_images.append(os.path.join(screenshot_path ,"visual" + str(idx) + ".png"))
                print("save_visual" + str(idx) + "_scr_alt_img")
            all_counts.append(len(list_target))

            if idx < len(acc_image) and idx > 0 :
                idx = idx - 1
                product = product_counts[idx]
                product_option_counts = acc_product_count[idx].find_elements(By.CSS_SELECTOR,"ul.accessories__product-option-list") # 제품 옵션 버튼 

                for i in range(0,product): # 제품 수 만큼 반복  
                    product_item = acc_product_count[idx].find_elements(By.CSS_SELECTOR,"li.accessories__product-item")
                    product_item_name = product_item[i].find_element(By.CSS_SELECTOR,"div.accessories__product-copy > h3").text # 제품 이름 추출

                    if product_option[idx] > 0:
                            product_option_count = product_option_counts[i].find_elements(By.CSS_SELECTOR,"li.accessories__product-option-item")
                                        
                    if click_count == len(product_color):
                        break
                                        
                    colorchip_click = product_color[click_count].find_elements(By.CSS_SELECTOR,"li.accessories__colorchip-item")
                        # 컬러칩 클릭할 요소 추출
                    count = len(colorchip_click)
                    for j in range(0,count):
                                            
                        if product_option[idx] > 0 and j >= (int(count/2)):
                            num = len(product_option_count) - 1
                            product_option_count[num].click()
                            product_name = product_option_count[num].find_element(By.CSS_SELECTOR,"button.accessories__product-option-btn").text # 제품 기종 추출

                        elif product_option[idx] == 0 :
                            product_name = ""
                        else : 
                            product_name = product_option_count[0].find_element(By.CSS_SELECTOR,"button.accessories__product-option-btn").text #제품 기종 추출

                        colorchip_click[j].click()
                        product_item_color = product_item[i].find_element(By.CSS_SELECTOR,"span.accessories__product-current").text # 제품 색 추출

                        a = product_color[color].find_element(By.TAG_NAME,"img")
                        list_target.append("data-src-pc") # 이미지 태그 이름
                        list_route.append(a.get_attribute("data-src-pc")) # 값
                        list_target.append("data-src-tablet") # 이미지 태그 이름
                        list_route.append(a.get_attribute("data-src-tablet"))# 값
                        list_target.append("data-src-mobile") # 이미지 태그 이름
                        list_route.append(a.get_attribute("data-src-mobile"))# 값
                        list_target.append("alt") # 이미지 태그 이름
                        list_route.append(a.get_attribute("alt"))# 값    
                        time.sleep(0.3)
                                            
                        all_counts.append(len(list_target))

                        time.sleep(0.3)
                        size = product_item[i].size
                        location = product_item[i].location

                        left = location['x']
                        top = location['y'] - 96
                        right = left + size['width']
                        bottom = size['height'] 
                                    
                        self.driver.execute_script(f"window.scrollTo({left},{top})")
                                    
                        time.sleep(0.5)

                        screenshot = self.driver.get_screenshot_as_png()
                        image = PILImage.open(BytesIO(screenshot))
                        image = image.crop((left, 96, right, bottom + 96)) 

                        image.save(os.path.join(screenshot_path, product_item_name +" "+ product_name +" "+ product_item_color + ".png"))
                        all_images.append(os.path.join(screenshot_path, product_item_name +" "+ product_name +" "+ product_item_color + ".png"))
                        img = img + 1

                    click_count = click_count + 1
                    color = color + 1

                time.sleep(0.3)
                counts = counts + 1
                # 구하고자 하는 엑서러리 페이지가 DM3 일경우
                if idx == 0:
                    # print("DM_1_2_save_cases_src_alt_img")
                    print("DM_3_save_cases_src_alt_img")
                elif idx == 1:
                    # print("DM_1_2_save_chargers_src_alt_img")
                    print("DM_3_save_pen_src_alt_img")
                elif idx == 2:
                    # print("DM 1/2_save_buds_src_alt_img")
                    print("DM_3_save_chargers_src_alt_img")
                elif idx == 3:
                    # print("DM 1/2_save_watch_src_alt_img")
                    print("DM_3_save_buds_src_alt_img")
                elif idx == 4:
                    # print("-")
                    print("DM_3_save_watch_src_alt_img")
                    bann = 0
                    acc_banner = self.driver.find_elements(By.CSS_SELECTOR,"div.common-banner__container")

                if idx == len(product_counts)-1:
                    for banner in acc_banner:
                        banner_img = banner.find_element(By.TAG_NAME,"img")
                        list_target.append("data-src-pc") # 이미지 태그 이름
                        list_route.append(banner_img.get_attribute("data-src-pc")) # 값
                        list_target.append("data-src-tablet") # 이미지 태그 이름
                        list_route.append(banner_img.get_attribute("data-src-tablet"))# 값
                        list_target.append("data-src-mobile") # 이미지 태그 이름
                        list_route.append(banner_img.get_attribute("data-src-mobile"))# 값
                        list_target.append("alt") # 이미지 태그 이름
                        list_route.append(banner_img.get_attribute("alt"))# 값

                        all_counts.append(len(list_target))
                        
                        banner_inner_img = self.driver.find_elements(By.CSS_SELECTOR,"div.common-banner__item-inner")  
                        time.sleep(0.3)
                        size = banner_inner_img[bann].size
                        location = banner_inner_img[bann].location

                        left = location['x']
                        top = location['y'] - 96
                        right = left + size['width']
                        bottom = size['height'] 
                            
                        self.driver.execute_script(f"window.scrollTo({left},{top})")
                            
                        time.sleep(0.5)

                        screenshot = self.driver.get_screenshot_as_png()
                        image = PILImage.open(BytesIO(screenshot))
                        image = image.crop((left, 96, right, bottom + 96)) 

                        image.save(os.path.join(screenshot_path,"Bottom_banner"+str(bann)+".png"))
                        all_images.append(os.path.join(screenshot_path,"Bottom_banner"+str(bann)+".png"))

                        bann = bann + 1
                        print("save_bottom_banner"+str(bann)+"_src_alt_img")
                        # inners = driver.find_elements(By.CSS_SELECTOR,"div.common-banner__text-inner")
                        if acc_banner[1] == banner:
                            icon = self.driver.find_element(By.CSS_SELECTOR,"div.common-banner__ar-icon")
                            icon = icon.find_element(By.TAG_NAME,"img")
                            list_target.append("data-src-pc") # 이미지 태그 이름
                            list_route.append(icon.get_attribute("data-src-pc")) # 값
                            list_target.append("alt") # 이미지 태그 이름
                            list_route.append(icon.get_attribute("alt"))# 값

                            qr = self.driver.find_element(By.CSS_SELECTOR,"div.common-banner__ar-qr")
                            qr = qr.find_element(By.TAG_NAME,"img")
                            list_target.append("data-src-pc") # 이미지 태그 이름
                            list_route.append(qr.get_attribute("data-src-pc")) # 값
                            list_target.append("alt") # 이미지 태그 이름
                            list_route.append(qr.get_attribute("alt"))# 값

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["number","대상", "이미지 경로","응답코드","이미지"])

        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 8
        ws.column_dimensions['C'].width = 80
        ws.column_dimensions['D'].width = 8
        ws.column_dimensions['E'].width = 80

        i = 0
        c = 1

        for i in range(0,len(list_target)):
            list_number.append(i+1)
            ws.append([list_number[i],list_target[i],list_route[i]])
                
            # 이미지 로드 requset 응답 추가
            if list_route[i] != "" and (list_route[i] is not None) :
                a = i + 1
                res = requests.get("https://www.samsung.com/" + list_route[i]) 
                time.sleep(0.5)
                if res.status_code == requests.codes.ok:
                    text = "Pass"
                    # 원하는 셀 선택
                    cell = ws.cell(row=a+1, column=4)
                    # 셀에 텍스트 추가
                    cell.value = text

            if c <= len(all_images) and i == (all_counts[c])-3:

                img = ExcelImage(all_images[c-1])
                time.sleep(0.3)
                ws.row_dimensions[i+1].height = 300

                if all_counts[c] == 7 or all_counts[c] == 11 or all_counts[c] == 175 or all_counts[c] == 199 or all_counts[c] == 235 or all_counts[c] == 279 or all_counts[c] == 283 or all_counts[c] == 291:
                    ws.row_dimensions[i+1].height = 100
                    img.width = 300
                    img.height = 100
                else : 
                    ws.row_dimensions[i+1].height = 300
                    img.width = 100
                    img.height = 300
                # 이미지를 첨부할 셀 지정
                cell = ws.cell(row=i+1, column=5)
                # 이미지를 셀에 첨부
                ws.add_image(img, cell.coordinate)
                c = c+1
        wb.save('accessories.xlsx')

        if len(list_number) == len(list_target):
            print("file create complete")

