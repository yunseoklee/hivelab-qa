from selenium import webdriver
from openpyxl.drawing.image import Image as ExcelImage
from selenium.webdriver.common.by import By
from selenium.common.exceptions import WebDriverException
import openpyxl, os, time

class BasePage(object):
    def __init__(self, driver):
        self.driver = driver

# 2번 산출물: 하이라이트 페이지 리포트
#메인 function 클래스
class MainFunction(BasePage):
    def page_report(self, excelName, sheetName, waitTime):
        
        # 엑셀 설정
        self.wb = openpyxl.load_workbook(filename=excelName)
        self.ws = self.wb[sheetName]

        # 스크린샷 저장을 위한 폴더 설정
        folder_name = "elements_screenshot"
        current_directory = os.getcwd()
        new_folder_path = os.path.join(current_directory, folder_name)
        if not os.path.exists(new_folder_path):
            os.makedirs(new_folder_path)
        else:
            pass

        self.ws['C3'] = self.driver.current_url

        # ID 속성이 비어있지 않은 모든 element를 찾고, 특정 ID들과 ID값들을 제외
        elements_with_id = self.driver.find_elements(By.XPATH,"//*[@id and not(@id='wrap') and not(@id='accessibility-navigation') and not(@id='header') and not(@id='gnb') and not(@id='g-products') and not(@id='g-campaigns') and not(@id='g-event') and not(@id='g-apps') and not(@id='cseSearchForm') and not(@id='cseOpenButton') and not(@id='addsearch-body') and not(@id='contents') and not(@id='subnav') and not(@id='faq_item1') and not(@id='faq_item2') and not(@id='faq_item3') and not(@id='faq_item4') and not(@id='faq_item5') and not(@id='faq_item6') and not(@id='faq_item7') and not(@id='faq_item8') and not(@id='desc-section') and not(@id='footer') and not(@id='accessibility-contrast') and not(@id='bandwidth-control') and not(@id='terms-and-conditions') and not(@id='footer-sitemap') and not(@id='modal-layer-popup') and not(@id='color-lavender') and not(@id='color-cream') and not(@id='color-phantom-black') and not(@id='color-green') and not(@id='')]")

        #번호 카운트를 위한 변수
        numberForCount = 0

        # ID를 가지고 있는 모든 element에 접근
        for element in elements_with_id:
            numberForCount = numberForCount + 1
            elementIdValue = element.get_attribute('id')
            # element ID값 엑셀에 쓰기
            self.ws["B"+str(numberForCount+4)] = elementIdValue

            # 해당 element로 이동
            self.driver.execute_script("arguments[0].scrollIntoView();", element)
            time.sleep(waitTime)
            # 100픽셀 만큼 스크롤업
            self.driver.execute_script("window.scrollBy(0,-100);")
            time.sleep(waitTime)
            # 스크린 샷 경로 및 이미지명 지정
            screenshot_path = os.path.join(new_folder_path, str(elementIdValue)+".png")
            # 스크린 샷
            try:
                element.screenshot(screenshot_path)
                elementImage = ExcelImage(screenshot_path)
                elementImage.height = 475
                elementImage.width = 600
                self.ws.add_image(elementImage, 'G'+str(numberForCount+4))
            except WebDriverException:
                self.ws["G"+str(numberForCount+4)] = "Can't take screenshot"
            
            # 클래스 이름이 blind인 element들을 각 element에서 찾기
            blind_elems = element.find_elements(By.CLASS_NAME, "blind")
            # 각 element text 변수
            element_text = element.text

            # 각 element에 이미지가 있는지 확인
            img_elems = element.find_elements(By.TAG_NAME, "img")
            if img_elems:
                all_values_str = ""
                all_alt_str = ""

                for i, img_elem in enumerate(img_elems):
                    img_src_pc = img_elem.get_attribute("data-src-pc")
                    img_src_tablet = img_elem.get_attribute("data-src-tablet")
                    img_src_mobile = img_elem.get_attribute("data-src-mobile")
                    img_alt = img_elem.get_attribute("alt")

                    values_list = []
                    alt_list = []

                    if img_src_pc:
                        values_list.append("Image src-pc: " + str(img_src_pc))
                    if img_src_tablet:
                        values_list.append("Image src-tablet: " + str(img_src_tablet))
                    if img_src_mobile:
                        values_list.append("Image src-mobile: " + str(img_src_mobile))
                    if img_alt:
                        alt_list.append("Image alt: " + str(img_alt))
                    
                    values_str = "\n".join(values_list)
                    all_values_str += values_str + "\n"
                    alt_str = "\n".join(alt_list)
                    all_alt_str += alt_str + "\n"
                
                if all_values_str == " ":
                    self.ws['C'+str(numberForCount+4)] = "None"
                else:
                    self.ws['C'+str(numberForCount+4)] = all_values_str.strip()
                if all_alt_str == " ":
                    self.ws['D'+str(numberForCount+4)] = "None"
                else:
                    self.ws['D'+str(numberForCount+4)] = all_alt_str.strip()
            
            if blind_elems:
                all_blind_str = ""

                for blind_elem in blind_elems:
                    blind_text = blind_elem.get_attribute("innerHTML")
                    blind_list = []
                    if blind_text:
                        blind_list.append("Blind text: " + str(blind_text))
                    
                    blinds_str = "\n".join(blind_list)
                    all_blind_str += blinds_str + "\n"

                    element_text = element_text.replace(blind_elem.text, "")

                if all_blind_str == " ":
                    self.ws['E'+str(numberForCount+4)] = "None"
                else:
                    self.ws['E'+str(numberForCount+4)] = all_blind_str.strip()
            
            if element_text == "":
                self.ws['F'+str(numberForCount+4)] = ""
            else:
                self.ws['F'+str(numberForCount+4)] = element_text
        # 파일저장
        self.wb.save(filename=excelName)