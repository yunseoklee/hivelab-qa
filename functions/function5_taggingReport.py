from selenium.webdriver.common.by import By
from openpyxl.drawing.image import Image as ExcelImage
from selenium.common.exceptions import InvalidSelectorException, NoSuchElementException
import openpyxl, os, time, re

class BasePage(object):
    def __init__(self, driver):
        self.driver = driver

#메인 function 클래스
class MainFunction(BasePage):
    # 5번 산출물: 태깅 자동화 리포트
    def tagging_automation_report(self, excelName, sheetName, waitTime):
        
        # 윈도우 사이즈 500x900으로 변경
        self.driver.set_window_size(500, 900)

        # 엑셀 설정
        self.wb = openpyxl.load_workbook(filename=excelName)
        self.ws = self.wb[sheetName]
        self.ws['E2'] = self.driver.current_url

        # 스크린샷 저장을 위한 폴더 설정
        folder_name = "tagging_screenshot"
        current_directory = os.getcwd()
        new_folder_path = os.path.join(current_directory, folder_name)
        if not os.path.exists(new_folder_path):
            os.makedirs(new_folder_path)
        else:
            pass

        time.sleep(waitTime)

        number = 0
        #페이지 전체 source 코드 가져오기
        html_source = self.driver.page_source

        for i in range(4, 303):
            # 엑셀에서 태깅 코드 가져오기
            excelTaggingCode = self.ws['C'+str(i)].value
            if excelTaggingCode is not None and excelTaggingCode in html_source:
                number = number + 1
                self.ws['E'+str(i)] = "PASS"

                # 태깅 형식 변경
                attribute_regex = r'(\S+)="([^"]*)"'
                matches = re.findall(attribute_regex, excelTaggingCode)
                desired_text = '[' + ']['.join([f'{attr}="{value}"' for attr, value in matches]) + ']'

                try:
                    # 해당 태깅의 element 찾기
                    element = self.driver.find_element(By.CSS_SELECTOR, f'{str(desired_text)}')

                    # 태깅 정보 엑셀에 쓰기
                    self.ws['D'+str(i)] = element.tag_name
                    self.ws['G'+str(i)] = str(element.is_displayed())
                    self.ws['F'+str(i)] = element.text

                    # element의 x 와 y 좌표 구하기
                    x = element.location['x'] - self.driver.execute_script("return window.innerWidth") / 2
                    y = element.location['y'] - self.driver.execute_script("return window.innerHeight") / 2
                    # element로 스크롤
                    self.driver.execute_script("window.scrollTo({0}, {1});".format(x, y))

                    # 해당 태깅 버튼에 빨간색 강조 처리
                    self.driver.execute_script("arguments[0].style.background = 'red'", element)

                    time.sleep(waitTime)

                    screenshot_path = os.path.join(new_folder_path, str(number)+".png")

                    if str(element.is_displayed()) == "True":
                        self.driver.save_screenshot(screenshot_path)
                        tagImage = ExcelImage(screenshot_path)
                        tagImage.height = 500
                        tagImage.width = 900
                        self.ws.add_image(tagImage, 'H' +str(i))
                    else:
                        pass

                    #강조처리 했던것 초기화
                    self.driver.execute_script("arguments[0].style.background = ''", element)
                    time.sleep(waitTime)
                except InvalidSelectorException as e:
                    print("InvalidSelector Exception occurred: {}".format(str(e)))
                except NoSuchElementException as e:
                    print("NoSuchElement Exception occurred: {}".format(str(e)))
            elif excelTaggingCode is not None and excelTaggingCode not in html_source:
                self.ws['E'+str(i)] = "FAIL"
            elif excelTaggingCode is None:
                self.ws['E'+str(i)] = "None"
            else:
                pass
        # 파일저장
        self.wb.save(filename=excelName)