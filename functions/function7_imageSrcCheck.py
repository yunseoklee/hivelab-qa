from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import time
from openpyxl import load_workbook
from bs4 import BeautifulSoup

class BasePage(object):
    def __init__(self, driver):
        self.driver = driver

#메인 function 클래스
class MainFunction(BasePage):
    # 7번 산출물: 이미지 경로 유효성 체크
    def imageSrcCheck(self):
        time.sleep(1)

        html = self.driver.page_source

        self.driver.quit()

        soup = BeautifulSoup( html, 'html.parser')

        wb = load_workbook("imgsrc_format.xlsx")
        ws = wb.active
        # with open("samsung.html", "w", encoding ="utf-8") as f : 
        #     f.write(res.text)#CP949
        kv_source = soup.findAll('source')
        content = soup.find('div',{'id':'contents'}).find_all("img")
        img_src = soup.findAll('img')

        temp=[]
        for i in kv_source :  
                srcset = i.get('srcset')
                if srcset is not None :
                    temp.append(['soucre',srcset])
        #print(temp)
        aa=[]
        for j in content : 
            pc = j.get('data-src-pc')
            tablet = j.get('data-src-tablet')
            mobile = j.get('data-src-mobile')
            alt = j.get('alt')
            if pc is not None :
                aa.append(['data-src-pc',pc])
            if tablet is not None :
                aa.append(['data-src-tablet',tablet])
            if mobile is not None :
                aa.append(['data-src-mobile',mobile])
            aa.append(['alt',alt])
        temp.extend(aa)
        sum=len(temp)
        print('총이미지:',sum)

        for x in range(2,len(temp)+2) :
            num=0
            ws['A'+str(x)] = x-1
            for y in range(2,4) : 
                ws.cell(row=x,column=y).value=temp[x-2][num]
                num=num+1

        wb.close()

        ex_export = input("저장할 파일 이름 : ")
        wb.save(ex_export+'.xlsx')

        # print("완료")